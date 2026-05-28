import os
import pandas as pd
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def normalize_passport_no(val):
    if pd.isna(val):
        return ""
    # Chuyển về chuỗi, viết hoa, bỏ khoảng trắng thừa
    return str(val).strip().upper()

def process_excel_and_matching(excel_path, ocr_results, output_excel_path):
    """
    Đọc Excel, đối chiếu Passport No. và xuất file kết quả Excel duy nhất gồm 2 Sheet:
    - Sheet 1: ATTENDANT REPORT (Dữ liệu gốc + cột Date of Issue ngoài cùng bên phải)
    - Sheet 2: Báo cáo Đối chiếu (Thống kê và chi tiết đối chiếu được thiết kế cực đẹp)
    """
    # 1. ĐỌC DỮ LIỆU
    df = pd.read_excel(excel_path)
    
    # Chuẩn hóa cột Document Number trong excel để đối chiếu
    df['normalized_doc_no'] = df['Document Number'].apply(normalize_passport_no)
    
    # Tạo từ điển map từ Passport No. (chuẩn hóa) sang Date of Issue
    passport_to_date = {}
    for res in ocr_results:
        p_no = normalize_passport_no(res.get("passport_no"))
        d_issue = res.get("date_of_issue")
        if p_no and d_issue:
            # Lưu ý: d_issue có định dạng YYYY/MM/DD
            passport_to_date[p_no] = d_issue

    # Tạo cột Date of Issue mới dựa trên đối chiếu
    df['Date of Issue'] = df['normalized_doc_no'].map(passport_to_date)
    
    # Loại bỏ cột tạm dùng để đối chiếu
    df.drop(columns=['normalized_doc_no'], inplace=True)
    
    # Di chuyển cột 'Date of Issue' ra sau cột 'Valid Until' (ngoài cùng bên phải)
    cols = list(df.columns)
    if 'Date of Issue' in cols:
        cols.remove('Date of Issue')
        cols.append('Date of Issue')
        df = df[cols]

    # 2. KHỞI TẠO WORKBOOK OPENPYXL
    wb = Workbook()
    
    # --- STYLE CHUNG ---
    font_family = "Arial"
    font_regular = Font(name=font_family, size=10)
    font_bold = Font(name=font_family, size=10, bold=True)
    font_header = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    font_title = Font(name=font_family, size=14, bold=True, color="1A365D")
    
    fill_header = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid") # Xanh Navy đậm
    fill_zebra = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid") # Xám nhạt
    fill_success_bg = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid") # Xanh lá nhạt
    font_success = Font(name=font_family, size=10, bold=True, color="137333")
    
    fill_fail_bg = PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid") # Đỏ nhạt
    font_fail = Font(name=font_family, size=10, bold=True, color="C2185B")
    
    border_thin = Side(style='thin', color='D9D9D9')
    border_double = Side(style='double', color='1A365D')
    
    box_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    bottom_double_border = Border(bottom=border_double)

    # --- SHEET 1: ATTENDANT REPORT ---
    ws1 = wb.active
    ws1.title = "ATTENDANT REPORT"
    ws1.views.sheetView[0].showGridLines = True
    
    # Ghi headers
    ws1.append(cols)
    for col_idx in range(1, len(cols) + 1):
        cell = ws1.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = box_border
    
    # Ghi dữ liệu
    for row_idx, row in enumerate(df.values, start=2):
        ws1.append(list(row))
        # Áp dụng format cho từng ô dữ liệu
        for col_idx in range(1, len(cols) + 1):
            cell = ws1.cell(row=row_idx, column=col_idx)
            cell.font = font_regular
            cell.border = box_border
            
            # Tô màu zebra cho các dòng chẵn để tăng tính thẩm mỹ
            if row_idx % 2 == 0:
                cell.fill = fill_zebra
                
            # Định dạng căn lề
            col_name = cols[col_idx - 1]
            if col_name in ['Index', 'Date', 'Time In', 'Time Out', 'Gender', 'Nationality', 'Valid Until', 'Date of Issue']:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_name in ['Document Number']:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                # Highlight nhẹ ô Date of Issue vừa điền thành công
                if col_name == 'Document Number':
                    doc_val = normalize_passport_no(cell.value)
                    if doc_val in passport_to_date:
                        ws1.cell(row=row_idx, column=len(cols)).fill = fill_success_bg
                        ws1.cell(row=row_idx, column=len(cols)).font = font_success
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Tự động căn chỉnh độ rộng cột cho Sheet 1
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # --- SHEET 2: Báo cáo Đối chiếu (Matching Report) ---
    ws2 = wb.create_sheet(title="Báo cáo Đối chiếu")
    ws2.views.sheetView[0].showGridLines = True
    
    # Tiêu đề báo cáo
    ws2.merge_cells("A2:F2")
    title_cell = ws2["A2"]
    title_cell.value = "BÁO CÁO KẾT QUẢ ĐỐI CHIẾU THÔNG TIN HỘ CHIẾU"
    title_cell.font = font_title
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    ws2.merge_cells("A3:F3")
    subtitle_cell = ws2["A3"]
    subtitle_cell.value = "Hệ thống Trí tuệ Nhân tạo OCR Tự động (Antigravity AI) | Xuất báo cáo Excel chuẩn"
    subtitle_cell.font = Font(name=font_family, size=10, italic=True, color="7F8C8D")
    subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Thông tin chung
    ws2["A5"] = "Ngày thực hiện:"
    ws2["A5"].font = font_bold
    ws2["B5"] = datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    ws2["B5"].font = font_regular
    
    ws2["A6"] = "Trạng thái:"
    ws2["A6"].font = font_bold
    ws2["B6"] = "Hoàn thành"
    ws2["B6"].font = font_success
    ws2["B6"].fill = fill_success_bg
    ws2["B6"].alignment = Alignment(horizontal="center")
    
    # Tính toán số liệu thống kê
    total_images = len(ocr_results)
    successful_ocr = sum(1 for r in ocr_results if r.get("passport_no") and r.get("date_of_issue"))
    matched_rows = df['Date of Issue'].notna().sum()
    total_excel_rows = len(df)
    
    # Bảng Thống kê tổng quan
    ws2["D5"] = "Tổng số ảnh quét:"
    ws2["D5"].font = font_bold
    ws2["E5"] = total_images
    ws2["E5"].font = font_regular
    ws2["E5"].alignment = Alignment(horizontal="left")
    
    ws2["D6"] = "OCR Thành công:"
    ws2["D6"].font = font_bold
    ws2["E6"] = f"{successful_ocr} / {total_images} ({successful_ocr/total_images*100:.1f}%)"
    ws2["E6"].font = font_regular
    
    ws2["D7"] = "Khớp dữ liệu Excel:"
    ws2["D7"].font = font_bold
    ws2["E7"] = f"{matched_rows} / {total_excel_rows} ({matched_rows/total_excel_rows*100:.1f}%)"
    ws2["E7"].font = font_regular

    # Tạo bảng chi tiết kết quả đối chiếu
    start_row = 9
    headers2 = ["STT", "Tên File Ảnh", "Số Passport Nhận Dạng", "Số Passport Trong Excel", "Ngày Cấp (Date of Issue)", "Kết Quả Đối Chiếu"]
    
    # Ghi headers cho bảng chi tiết
    for col_idx, header in enumerate(headers2, start=1):
        cell = ws2.cell(row=start_row, column=col_idx)
        cell.value = header
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = box_border
        
    ws2.row_dimensions[start_row].height = 25
    
    # Bản đồ tra cứu nhanh OCR
    ocr_map = {normalize_passport_no(r.get("passport_no")): r for r in ocr_results if r.get("passport_no")}
    
    # Điền dữ liệu bảng chi tiết
    stt = 1
    for idx, row in df.iterrows():
        current_row = start_row + stt
        ws2.row_dimensions[current_row].height = 20
        
        doc_no = str(row['Document Number'])
        normalized_doc = normalize_passport_no(doc_no)
        d_issue = row.get('Date of Issue')
        
        file_name = "-"
        scanned_no = "-"
        if normalized_doc in ocr_map:
            file_name = ocr_map[normalized_doc].get("file_name", "-")
            scanned_no = ocr_map[normalized_doc].get("passport_no", "-")
            
        status_text = "Khớp thành công" if pd.notna(d_issue) else "Không tìm thấy ảnh"
        
        # Ghi các ô
        ws2.cell(row=current_row, column=1, value=stt)
        ws2.cell(row=current_row, column=2, value=file_name)
        ws2.cell(row=current_row, column=3, value=scanned_no)
        ws2.cell(row=current_row, column=4, value=doc_no if pd.notna(doc_no) else "-")
        ws2.cell(row=current_row, column=5, value=str(d_issue) if pd.notna(d_issue) else "-")
        ws2.cell(row=current_row, column=6, value=status_text)
        
        # Format style các ô trong bảng
        for c in range(1, 7):
            cell = ws2.cell(row=current_row, column=c)
            cell.font = font_regular
            cell.border = box_border
            if current_row % 2 == 0:
                cell.fill = fill_zebra
            
            # Căn lề
            if c in [1, 3, 4, 5, 6]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
        # Tô màu trạng thái cột Kết Quả
        status_cell = ws2.cell(row=current_row, column=6)
        if pd.notna(d_issue):
            status_cell.fill = fill_success_bg
            status_cell.font = font_success
        else:
            status_cell.fill = fill_fail_bg
            status_cell.font = font_fail
            
        stt += 1

    # Tự động căn chỉnh cột cho Sheet 2
    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        # Bỏ qua dòng tiêu đề dòng 2 và 3 khi tính độ rộng
        ws2.column_dimensions[col_letter].width = max(max_len + 3, 14)
        
    # Thiết lập độ rộng cột A dòng STT nhỏ lại
    ws2.column_dimensions['A'].width = 8

    # 3. LƯU WORKBOOK GỘP
    os.makedirs(os.path.dirname(output_excel_path), exist_ok=True)
    wb.save(output_excel_path)
    print(f"Đã xuất file excel báo cáo tích hợp 2 Sheet thành công: {output_excel_path}")
    return df

def generate_word_report(df_matched, ocr_results, output_docx_path):
    """
    (Giữ lại hàm để tránh lỗi import nhưng không dùng nữa vì đã gộp đầu ra vào Excel)
    """
    pass
