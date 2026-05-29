import pandas as pd
import os
import datetime
import re


def normalize_passport_no(val):
    """Chuẩn hóa số passport: uppercase, trim."""
    if pd.isna(val) or val is None:
        return ""
    return str(val).strip().upper()


def normalize_character_variants(val):
    """Chuẩn hóa các ký tự dễ nhầm lẫn do lỗi OCR (O->0, I->1, etc.) để so khớp."""
    if not val:
        return ""
    val = str(val).strip().upper()
    replacements = {
        'O': '0',
        'I': '1',
        'L': '1',
        'S': '5',
        'Z': '2',
        'G': '6',
        'B': '8'
    }
    for char, rep in replacements.items():
        val = val.replace(char, rep)
    return val


def ocr_date_to_excel_int(date_str):
    """
    Chuyển đổi ngày từ format OCR (DD/MM/YYYY) sang format Excel (YYYYMMDD int).
    Trả về int hoặc None nếu không parse được.
    """
    if not date_str or not isinstance(date_str, str):
        return None
    date_str = date_str.strip()

    # Try DD/MM/YYYY or DD-MM-YYYY
    m = re.match(r"(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{4})", date_str)
    if m:
        d, mo, y = m.groups()
        try:
            return int(f"{y}{mo.zfill(2)}{d.zfill(2)}")
        except ValueError:
            return None

    # Try YYYY/MM/DD or YYYY-MM-DD
    m = re.match(r"(\d{4})[/\-.](\d{1,2})[/\-.](\d{1,2})", date_str)
    if m:
        y, mo, d = m.groups()
        try:
            return int(f"{y}{mo.zfill(2)}{d.zfill(2)}")
        except ValueError:
            return None

    # Try YYYYMMDD already
    m = re.match(r"^(\d{8})$", date_str)
    if m:
        try:
            return int(date_str)
        except ValueError:
            return None

    return None


def excel_int_to_display(val):
    """Chuyển YYYYMMDD (int/float) sang DD/MM/YYYY để hiển thị."""
    if pd.isna(val) or val is None:
        return ""
    s = str(int(float(val)))
    if len(s) == 8:
        return f"{s[6:8]}/{s[4:6]}/{s[0:4]}"
    return s


class ExcelMatcher:
    """
    Đọc file Excel SecureScan, hỗ trợ:
    - Tìm dòng theo Document Number (passport ID)
    - Điền thông tin còn thiếu từ OCR
    - Thêm Date of Issue
    """

    def __init__(self, excel_path):
        self.excel_path = excel_path
        self.df = None
        self.matched_passports = {}  # passport_no -> ocr_data for report
        self.unmatched_passports = []  # OCR results not found in Excel
        self.load_data()

    def load_data(self):
        """Đọc file Excel SecureScan."""
        try:
            if self.excel_path and os.path.exists(self.excel_path):
                self.df = pd.read_excel(self.excel_path)
                # Đảm bảo cột Document Number tồn tại
                if 'Document Number' not in self.df.columns:
                    print("WARNING: Cột 'Document Number' không tồn tại trong file Excel!")
                    self.df = None
        except Exception as e:
            print(f"Lỗi khi đọc file Excel: {e}")
            self.df = None

    def find_row_by_passport(self, passport_num):
        """
        Tìm index dòng trong DataFrame theo Document Number.
        Hỗ trợ so khớp chính xác và khớp mờ (Fuzzy matching) phòng trường hợp OCR sai ký tự.
        """
        if self.df is None or self.df.empty or not passport_num:
            return None

        search_term = normalize_passport_no(passport_num)
        if not search_term:
            return None

        # 1. So khớp chính xác 100% trước
        for idx, row in self.df.iterrows():
            doc_no = normalize_passport_no(row.get('Document Number'))
            if doc_no == search_term:
                return idx

        # 2. Khớp theo chuẩn hóa biến thể ký tự tương đồng (O/0, I/1, ...)
        norm_search = normalize_character_variants(search_term)
        for idx, row in self.df.iterrows():
            doc_no = normalize_passport_no(row.get('Document Number'))
            if normalize_character_variants(doc_no) == norm_search:
                return idx

        # 3. Quét khớp mờ sử dụng SequenceMatcher (độ tương đồng >= 80%)
        import difflib
        best_match_idx = None
        best_score = 0.0
        
        for idx, row in self.df.iterrows():
            doc_no = normalize_passport_no(row.get('Document Number'))
            if not doc_no:
                continue
            # Chỉ so khớp nếu độ dài tương đồng để tránh khớp nhầm hoàn toàn
            if abs(len(doc_no) - len(search_term)) <= 2:
                score = difflib.SequenceMatcher(None, search_term, doc_no).ratio()
                if score > best_score:
                    best_score = score
                    best_match_idx = idx

        # Ngưỡng chấp nhận khớp mờ là 80% trở lên
        if best_score >= 0.8:
            return best_match_idx

        return None

    def fill_missing_data(self, row_index, ocr_data):
        """
        Điền thông tin còn thiếu vào dòng Excel từ kết quả OCR.
        Chỉ điền nếu trường đó đang trống/NaN.
        
        ocr_data: dict với các key từ ocr_service (so_passport, ho_ten, ngay_sinh, ...)
        """
        if self.df is None or row_index is None:
            return

        row = self.df.loc[row_index]

        # Tách họ tên từ OCR
        full_name = str(ocr_data.get("ho_ten", "")).strip().upper()
        parts = full_name.split(" ", 1)
        ocr_last_name = parts[0] if len(parts) >= 1 else ""
        ocr_first_name = parts[1] if len(parts) >= 2 else ""

        # Gender: chuyển từ "Nam"/"Nữ" sang "M"/"F"
        gender_raw = str(ocr_data.get("gioi_tinh", "")).strip().upper()
        if "NAM" in gender_raw or gender_raw == "M":
            ocr_gender = "M"
        elif "NỮ" in gender_raw.upper() or "NU" in gender_raw or gender_raw == "F":
            ocr_gender = "F"
        else:
            ocr_gender = gender_raw

        # Date of Birth: DD/MM/YYYY -> YYYYMMDD int
        ocr_dob = ocr_date_to_excel_int(ocr_data.get("ngay_sinh", ""))

        # Valid Until: DD/MM/YYYY -> YYYYMMDD float
        ocr_valid_until = ocr_date_to_excel_int(ocr_data.get("ngay_het_han", ""))

        # Nationality
        ocr_nationality = str(ocr_data.get("quoc_tich", "")).strip().upper()

        # Document Number
        ocr_doc_num = str(ocr_data.get("so_passport", "")).strip().upper()

        # Điền các trường thiếu
        field_mapping = {
            'First Name': ocr_first_name,
            'Last Name': ocr_last_name,
            'Gender': ocr_gender,
            'Nationality': ocr_nationality,
            'Document Number': ocr_doc_num,
        }

        for col, ocr_val in field_mapping.items():
            if col in self.df.columns and ocr_val:
                current_val = row.get(col)
                if pd.isna(current_val) or str(current_val).strip() == "":
                    self.df.at[row_index, col] = ocr_val

        # Date of Birth (int)
        if 'Date of Birth' in self.df.columns and ocr_dob is not None:
            current_dob = row.get('Date of Birth')
            if pd.isna(current_dob):
                self.df.at[row_index, 'Date of Birth'] = ocr_dob

        # Valid Until (float)
        if 'Valid Until' in self.df.columns and ocr_valid_until is not None:
            current_vu = row.get('Valid Until')
            if pd.isna(current_vu):
                self.df.at[row_index, 'Valid Until'] = float(ocr_valid_until)

    def add_date_of_issue(self, row_index, date_value):
        """
        Thêm Date of Issue vào cột cuối cho dòng chỉ định.
        date_value: string DD/MM/YYYY từ OCR → chuyển thành YYYYMMDD int.
        """
        if self.df is None or row_index is None:
            return

        # Đảm bảo cột Date of Issue tồn tại
        if 'Date of Issue' not in self.df.columns:
            self.df['Date of Issue'] = pd.NA

        doi_int = ocr_date_to_excel_int(date_value)
        if doi_int is not None:
            self.df.at[row_index, 'Date of Issue'] = doi_int

    def process_ocr_result(self, ocr_data, file_name=""):
        """
        Xử lý một kết quả OCR: tìm dòng trùng → điền thiếu → thêm Date of Issue.
        Trả về trạng thái đối chiếu (string).
        """
        passport_num = ocr_data.get("so_passport", "")
        if not passport_num:
            return "Không có số Passport"

        row_index = self.find_row_by_passport(passport_num)

        if row_index is not None:
            # Trùng: điền thông tin thiếu
            self.fill_missing_data(row_index, ocr_data)
            # Thêm Date of Issue
            date_of_issue = ocr_data.get("ngay_cap", "")
            if date_of_issue:
                self.add_date_of_issue(row_index, date_of_issue)

            self.matched_passports[passport_num] = {
                "file_name": file_name,
                "ocr_data": ocr_data,
                "row_index": row_index
            }
            return "Khớp Excel - Đã điền thông tin"
        else:
            # Không tìm thấy trong Excel
            self.unmatched_passports.append({
                "file_name": file_name,
                "passport_num": passport_num,
                "ocr_data": ocr_data
            })
            return "Không có trong Excel"

    def get_match_stats(self):
        """Trả về thống kê đối chiếu."""
        total_excel = len(self.df) if self.df is not None else 0
        matched = len(self.matched_passports)
        unmatched = len(self.unmatched_passports)
        return {
            "total_excel": total_excel,
            "matched": matched,
            "unmatched_ocr": unmatched
        }


def _sanitize_value(val):
    """Convert NaN/pd.NA to None for openpyxl compatibility."""
    if val is pd.NA:
        return None
    try:
        if pd.isna(val):
            return None
    except (TypeError, ValueError):
        pass
    return val


def save_updated_excel(matcher, output_folder):
    """
    Lưu DataFrame đã cập nhật ra file Excel mới, format giống mau.xls.
    Tạo 2 sheets: ATTENDANT REPORT + Báo cáo Đối chiếu.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if matcher.df is None:
        return "Không có dữ liệu Excel", False

    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = os.path.join(output_folder, f"SecureScan_Updated_{timestamp}.xlsx")

    df = matcher.df.copy()

    # Đảm bảo cột Date of Issue ở cuối
    cols = list(df.columns)
    if 'Date of Issue' in cols:
        cols.remove('Date of Issue')
        cols.append('Date of Issue')
        df = df[cols]

    # --- STYLE CHUNG ---
    font_family = "Arial"
    font_regular = Font(name=font_family, size=10)
    font_bold = Font(name=font_family, size=10, bold=True)
    font_header = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    font_title = Font(name=font_family, size=14, bold=True, color="1A365D")

    fill_header = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    fill_zebra = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")
    fill_success_bg = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
    font_success = Font(name=font_family, size=10, bold=True, color="137333")
    fill_fail_bg = PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid")
    font_fail = Font(name=font_family, size=10, bold=True, color="C2185B")
    fill_filled_bg = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    font_filled = Font(name=font_family, size=10, color="E65100")

    border_thin = Side(style='thin', color='D9D9D9')
    box_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

    wb = Workbook()

    # ─── SHEET 1: ATTENDANT REPORT ──────────────────────────────
    ws1 = wb.active
    ws1.title = "ATTENDANT REPORT"

    # Headers
    ws1.append(cols)
    for col_idx in range(1, len(cols) + 1):
        cell = ws1.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = box_border

    # Data rows
    center_cols = {'Index', 'Date', 'Time', 'Time out', 'Gender', 'Nationality',
                   'Valid Until', 'Date of Issue', 'Document Number', 'Date of Birth'}

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        values = [_sanitize_value(row.get(c)) for c in cols]
        ws1.append(values)

        for col_idx in range(1, len(cols) + 1):
            cell = ws1.cell(row=row_idx, column=col_idx)
            cell.font = font_regular
            cell.border = box_border

            if row_idx % 2 == 0:
                cell.fill = fill_zebra

            col_name = cols[col_idx - 1]
            if col_name in center_cols:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # Highlight Date of Issue cell nếu đã được điền
        if 'Date of Issue' in cols:
            doi_col_idx = cols.index('Date of Issue') + 1
            doi_cell = ws1.cell(row=row_idx, column=doi_col_idx)
            if doi_cell.value is not None:
                doi_cell.fill = fill_success_bg
                doi_cell.font = font_success

    # Auto-width
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # ─── SHEET 2: Báo cáo Đối chiếu ───────────────────────────
    ws2 = wb.create_sheet(title="Báo cáo Đối chiếu")

    # Tiêu đề
    ws2.merge_cells("A2:F2")
    title_cell = ws2["A2"]
    title_cell.value = "BÁO CÁO KẾT QUẢ ĐỐI CHIẾU THÔNG TIN HỘ CHIẾU"
    title_cell.font = font_title
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws2.merge_cells("A3:F3")
    subtitle_cell = ws2["A3"]
    subtitle_cell.value = "Hệ thống OCR Hộ chiếu & Đối chiếu Dữ liệu Tự động"
    subtitle_cell.font = Font(name=font_family, size=10, italic=True, color="7F8C8D")
    subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Thống kê
    stats = matcher.get_match_stats()
    ws2["A5"] = "Ngày thực hiện:"
    ws2["A5"].font = font_bold
    ws2["B5"] = datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    ws2["B5"].font = font_regular

    ws2["A6"] = "Trạng thái:"
    ws2["A6"].font = font_bold
    ws2["B6"] = "Hoàn thành"
    ws2["B6"].font = font_success
    ws2["B6"].fill = fill_success_bg

    ws2["D5"] = "Tổng dòng Excel:"
    ws2["D5"].font = font_bold
    ws2["E5"] = stats["total_excel"]
    ws2["E5"].font = font_regular

    ws2["D6"] = "Khớp thành công:"
    ws2["D6"].font = font_bold
    total_ex = stats["total_excel"] if stats["total_excel"] > 0 else 1
    ws2["E6"] = f"{stats['matched']} / {stats['total_excel']} ({stats['matched']/total_ex*100:.1f}%)"
    ws2["E6"].font = font_regular

    ws2["D7"] = "Ảnh không trong Excel:"
    ws2["D7"].font = font_bold
    ws2["E7"] = stats["unmatched_ocr"]
    ws2["E7"].font = font_regular

    # Bảng chi tiết
    start_row = 9
    headers2 = ["STT", "Tên File Ảnh", "Số Passport OCR", "Số Passport Excel",
                "Date of Issue", "Kết Quả Đối Chiếu"]

    for col_idx, header in enumerate(headers2, start=1):
        cell = ws2.cell(row=start_row, column=col_idx)
        cell.value = header
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = box_border

    ws2.row_dimensions[start_row].height = 25

    stt = 1
    # Matched passports
    for p_no, info in matcher.matched_passports.items():
        current_row = start_row + stt
        ws2.row_dimensions[current_row].height = 20

        doi_val = info["ocr_data"].get("ngay_cap", "-")

        ws2.cell(row=current_row, column=1, value=stt)
        ws2.cell(row=current_row, column=2, value=info["file_name"])
        ws2.cell(row=current_row, column=3, value=p_no)
        ws2.cell(row=current_row, column=4, value=p_no)
        ws2.cell(row=current_row, column=5, value=doi_val if doi_val else "-")
        ws2.cell(row=current_row, column=6, value="Khớp thành công")

        for c in range(1, 7):
            cell = ws2.cell(row=current_row, column=c)
            cell.font = font_regular
            cell.border = box_border
            if current_row % 2 == 0:
                cell.fill = fill_zebra
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws2.cell(row=current_row, column=6).fill = fill_success_bg
        ws2.cell(row=current_row, column=6).font = font_success
        stt += 1

    # Unmatched passports
    for info in matcher.unmatched_passports:
        current_row = start_row + stt
        ws2.row_dimensions[current_row].height = 20

        ws2.cell(row=current_row, column=1, value=stt)
        ws2.cell(row=current_row, column=2, value=info["file_name"])
        ws2.cell(row=current_row, column=3, value=info["passport_num"])
        ws2.cell(row=current_row, column=4, value="-")
        ws2.cell(row=current_row, column=5, value="-")
        ws2.cell(row=current_row, column=6, value="Không có trong Excel")

        for c in range(1, 7):
            cell = ws2.cell(row=current_row, column=c)
            cell.font = font_regular
            cell.border = box_border
            if current_row % 2 == 0:
                cell.fill = fill_zebra
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws2.cell(row=current_row, column=6).fill = fill_fail_bg
        ws2.cell(row=current_row, column=6).font = font_fail
        stt += 1

    # Auto-width cho Sheet 2
    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws2.column_dimensions[col_letter].width = max(max_len + 3, 14)
    ws2.column_dimensions['A'].width = 8

    try:
        wb.save(output_filename)
        return output_filename, True
    except Exception as e:
        print(f"Error saving output Excel: {e}")
        return str(e), False


# ─── Legacy function (kept for backward compatibility) ──────────
def format_date(date_str, output_format="DD/MM/YYYY"):
    date_str = str(date_str).strip()
    if not date_str:
        return ""

    m = re.match(r"(\d{2})[-/](\d{2})[-/](\d{4})", date_str)
    if m:
        d, m_m, y = m.groups()
        if output_format == "DD/MM/YYYY":
            return f"{d}/{m_m}/{y}"
        elif output_format == "YYYYMMDD":
            return f"{y}{m_m}{d}"
        else:
            return f"{y}/{m_m}/{d}"

    m = re.match(r"(\d{4})[-/](\d{2})[-/](\d{2})", date_str)
    if m:
        y, m_m, d = m.groups()
        if output_format == "DD/MM/YYYY":
            return f"{d}/{m_m}/{y}"
        elif output_format == "YYYYMMDD":
            return f"{y}{m_m}{d}"
        else:
            return f"{y}/{m_m}/{d}"

    m = re.match(r"(\d{4})(\d{2})(\d{2})", date_str)
    if m:
        y, m_m, d = m.groups()
        if output_format == "DD/MM/YYYY":
            return f"{d}/{m_m}/{y}"
        elif output_format == "YYYYMMDD":
            return f"{y}{m_m}{d}"
        else:
            return f"{y}/{m_m}/{d}"

    return date_str
